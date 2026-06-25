import sqlite3
from datetime import datetime, timezone
from config import DB_PATH, EXCEL_PATH


def _conn():
    return sqlite3.connect(DB_PATH)


def _now():
    return datetime.now(timezone.utc).isoformat()


def init_db():
    with _conn() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS seen_jobs (
                job_id     TEXT PRIMARY KEY,
                source     TEXT,
                first_seen TIMESTAMP,
                title      TEXT,
                company    TEXT,
                url        TEXT
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS scored_jobs (
                job_id      TEXT PRIMARY KEY,
                score       INTEGER,
                title       TEXT,
                company     TEXT,
                url         TEXT,
                location    TEXT,
                sponsorship TEXT,
                source      TEXT,
                posted_at   TEXT,
                best_resume TEXT,
                ats_score   INTEGER,
                scored_at   TIMESTAMP,
                notified    INTEGER DEFAULT 0,
                applied     INTEGER DEFAULT 0,
                easy_apply  INTEGER DEFAULT 0
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS applied_jobs (
                job_id     TEXT PRIMARY KEY,
                applied_at TIMESTAMP,
                platform   TEXT,
                notes      TEXT
            )
        """)
        # Migrations for existing DBs
        for col, defn in [
            ("easy_apply",  "INTEGER DEFAULT 0"),
            ("source",      "TEXT"),
            ("posted_at",   "TEXT"),
            ("best_resume", "TEXT"),
            ("ats_score",   "INTEGER"),
            ("yoe_flag",    "TEXT DEFAULT ''"),
            ("reposted",    "INTEGER DEFAULT 0"),
        ]:
            try:
                con.execute(f"ALTER TABLE scored_jobs ADD COLUMN {col} {defn}")
            except sqlite3.OperationalError:
                pass
        con.commit()


def is_seen(job_id: str) -> bool:
    with _conn() as con:
        return con.execute(
            "SELECT 1 FROM seen_jobs WHERE job_id=?", (job_id,)
        ).fetchone() is not None


def mark_seen(job_id: str, source: str, title: str, company: str, url: str):
    with _conn() as con:
        con.execute(
            "INSERT OR IGNORE INTO seen_jobs "
            "(job_id, source, first_seen, title, company, url) VALUES (?,?,?,?,?,?)",
            (job_id, source, _now(), title, company, url),
        )
        con.commit()


def get_first_seen(job_id: str) -> str | None:
    with _conn() as con:
        row = con.execute(
            "SELECT first_seen FROM seen_jobs WHERE job_id=?", (job_id,)
        ).fetchone()
    return row[0] if row else None


def save_scored_job(job_id: str, score: int, title: str, company: str, url: str,
                    location: str, sponsorship: str, source: str = "",
                    posted_at: str | None = None, best_resume: str = "",
                    ats_score: int = 0, easy_apply: int = 0, yoe_flag: str = "",
                    reposted: int = 0):
    with _conn() as con:
        # UPSERT: insert with notified=0 for new jobs.
        # ON CONFLICT: update all fields EXCEPT notified — preserves notified=1
        # so already-emailed jobs are never re-sent even if the DB is rebuilt.
        con.execute(
            """INSERT INTO scored_jobs
               (job_id, score, title, company, url, location, sponsorship,
                source, posted_at, best_resume, ats_score,
                scored_at, notified, applied, easy_apply, yoe_flag, reposted)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,0,0,?,?,?)
               ON CONFLICT(job_id) DO UPDATE SET
                 score=excluded.score, title=excluded.title,
                 company=excluded.company, url=excluded.url,
                 location=excluded.location, sponsorship=excluded.sponsorship,
                 source=excluded.source, posted_at=excluded.posted_at,
                 best_resume=excluded.best_resume, ats_score=excluded.ats_score,
                 scored_at=excluded.scored_at, easy_apply=excluded.easy_apply,
                 yoe_flag=excluded.yoe_flag, reposted=excluded.reposted""",
            (job_id, score, title, company, url, location, sponsorship,
             source, posted_at, best_resume, ats_score, _now(), easy_apply, yoe_flag, reposted),
        )
        con.commit()


def mark_notified(job_id: str):
    with _conn() as con:
        con.execute("UPDATE scored_jobs SET notified=1 WHERE job_id=?", (job_id,))
        con.commit()


def mark_applied(job_id: str, platform: str, notes: str):
    with _conn() as con:
        con.execute(
            "INSERT OR REPLACE INTO applied_jobs "
            "(job_id, applied_at, platform, notes) VALUES (?,?,?,?)",
            (job_id, _now(), platform, notes),
        )
        con.execute("UPDATE scored_jobs SET applied=1 WHERE job_id=?", (job_id,))
        con.commit()
    _sync_excel()


def get_unnotified_scored_jobs() -> list:
    with _conn() as con:
        rows = con.execute(
            """SELECT job_id, score, title, company, url, location, sponsorship,
                      source, posted_at, best_resume, ats_score, easy_apply, yoe_flag, reposted
               FROM scored_jobs WHERE notified=0 ORDER BY score DESC"""
        ).fetchall()
    return [
        {
            "job_id": r[0], "score": r[1], "title": r[2], "company": r[3],
            "url": r[4], "location": r[5], "sponsorship": r[6],
            "source": r[7], "posted_at": r[8], "best_resume": r[9],
            "ats_score": r[10], "easy_apply": r[11], "yoe_flag": r[12] or "",
            "reposted": r[13] or 0,
        }
        for r in rows
    ]


def _sync_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        with _conn() as con:
            rows = con.execute("""
                SELECT sj.job_id, sj.title, sj.company, sj.url, sj.score,
                       sj.sponsorship, sj.location, sj.source, sj.posted_at,
                       sj.best_resume, sj.ats_score, sj.easy_apply,
                       sj.applied, aj.applied_at, aj.platform, aj.notes
                FROM scored_jobs sj
                LEFT JOIN applied_jobs aj ON sj.job_id = aj.job_id
                ORDER BY sj.score DESC
            """).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Applications"

        headers = [
            "Job ID", "Title", "Company", "URL", "Score", "Sponsorship",
            "Location", "Source", "Posted Date", "Best Resume", "ATS %",
            "Easy Apply", "Applied", "Applied At", "Platform", "Notes",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            score = row[4]
            color = "00B050" if score >= 90 else ("FF8C00" if score >= 80 else "000000")
            ws.cell(row=row_idx, column=5).font = Font(color=color, bold=True)

        col_widths = [15, 35, 22, 50, 7, 13, 22, 14, 20, 28, 7, 10, 8, 20, 14, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(EXCEL_PATH)
    except Exception as e:
        _log_error(f"Excel sync failed: {e}")


def rebuild_excel():
    _sync_excel()


def _log_error(msg: str):
    from config import ERROR_LOG
    with open(ERROR_LOG, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now(timezone.utc).isoformat()}] database.py: {msg}\n")
