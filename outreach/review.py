"""
Review queue — the human approval surface.

Two outputs from the current 'draft' rows:
  1. outreach_review.xlsx  — the EDITABLE surface. You can edit the Subject /
     Message cells and set the Action column to 'send' or 'skip'.
  2. outreach_queue.html   — a read-only visual preview to eyeball quickly.

The sender (outreach.py --send) reads outreach_review.xlsx back so your edits
and Action choices are honoured.
"""
import os
import html as _html

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from outreach.outreach_db import get_drafts, get_pending, get_all_sent

XLSX_PATH = r"C:\JobRadar\outreach_review.xlsx"
HTML_PATH = r"C:\JobRadar\outreach_queue.html"

PENDING_SHEET = "Pending"
HISTORY_SHEET = "Sent History"

# Column order in the Pending sheet. Editable cells: Subject, Message, Action.
COLS = ["Log ID", "Name", "Title", "Kind", "Company", "Role", "Channel",
        "Contact", "Subject", "Message", "Action", "Status"]
# History sheet columns (read-only audit trail).
HIST_COLS = ["Name", "Title", "Kind", "Company", "Role", "Channel",
             "Contact", "Status", "When"]
_ACTION_COL = COLS.index("Action") + 1   # 1-based
_MSG_COL    = COLS.index("Message") + 1
_SUBJ_COL   = COLS.index("Subject") + 1
_ID_COL     = COLS.index("Log ID") + 1


def _contact(d: dict) -> str:
    return d.get("email") or d.get("linkedin_url") or ""


def write_review(company: str | None = None, role: str | None = None) -> dict:
    """Write the tracker xlsx (Pending + Sent History across ALL companies) and
    an HTML preview scoped to this company/role batch."""
    pending = get_pending()
    history = get_all_sent()
    _write_xlsx(pending, history)
    # HTML preview shows just the batch you just built (drafts + li_pending for it).
    batch = [d for d in pending if (not company or d["company"] == company)
             and (not role or d["role"] == role)]
    _write_html(batch, company, role)
    return {"pending": len(pending), "history": len(history),
            "xlsx": XLSX_PATH, "html": HTML_PATH}


def _write_xlsx(pending: list[dict], history: list[dict]):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Pending ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = PENDING_SHEET
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    sep_fill = PatternFill("solid", fgColor="dbeafe")
    for c, h in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    r = 2
    last_company = None
    for d in pending:
        if d["company"] != last_company:
            # company separator row
            sc = ws.cell(row=r, column=1, value=f"▼ {d['company']}")
            sc.font = Font(bold=True, color="1e3a8a")
            for c in range(1, len(COLS) + 1):
                ws.cell(row=r, column=c).fill = sep_fill
            r += 1
            last_company = d["company"]
        row = [d["id"], d["full_name"], d["title"], d["kind"], d["company"],
               d["role"], d["channel"], _contact(d), d["subject"], d["message"],
               "", d["status"]]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == _MSG_COL))
        ws.cell(row=r, column=7).font = Font(
            bold=True, color="0a66c2" if d["channel"] == "linkedin" else "166534")
        r += 1

    widths = [7, 22, 28, 11, 16, 20, 10, 30, 28, 64, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet 2: Sent History (read-only audit trail) ───────────────────────
    hs = wb.create_sheet(HISTORY_SHEET)
    hhf = PatternFill("solid", fgColor="374151")
    for c, h in enumerate(HIST_COLS, 1):
        cell = hs.cell(row=1, column=c, value=h)
        cell.fill = hhf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    status_color = {"draft_created": "166534", "sent": "00a651",
                    "skipped": "9ca3af", "approved": "0a66c2"}
    for r2, d in enumerate(history, 2):
        when = (d.get("sent_at") or "")[:16].replace("T", " ")
        row = [d["full_name"], d["title"], d["kind"], d["company"], d["role"],
               d["channel"], _contact(d), d["status"], when]
        for c, val in enumerate(row, 1):
            hs.cell(row=r2, column=c, value=val)
        hs.cell(row=r2, column=8).font = Font(
            bold=True, color=status_color.get(d["status"], "374151"))
    hwidths = [22, 30, 11, 16, 20, 10, 30, 14, 18]
    for i, w in enumerate(hwidths, 1):
        hs.column_dimensions[get_column_letter(i)].width = w
    hs.freeze_panes = "A2"

    _safe_save(wb)


def _safe_save(wb):
    """Save the tracker. If it's open in Excel (locked), save to a timestamped
    fallback so data is never lost, and tell the user."""
    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        from datetime import datetime
        alt = XLSX_PATH.replace(".xlsx", f"_{datetime.now():%H%M%S}.xlsx")
        wb.save(alt)
        print(f"  ⚠ {XLSX_PATH} is open in Excel — saved a copy to:\n     {alt}\n"
              f"     (close the tracker and re-run to update the main file)")


def _badge(text: str, fg: str, bg: str) -> str:
    return (f'<span style="background:{bg};color:{fg};padding:2px 8px;border-radius:10px;'
            f'font-size:11px;font-weight:700;">{_html.escape(text)}</span>')


def _card(d: dict) -> str:
    ch = d["channel"]
    kind = d["kind"]
    ch_badge = _badge("✉ Email" if ch == "email" else "in LinkedIn",
                      "#fff", "#166534" if ch == "email" else "#0a66c2")
    kind_colors = {"recruiter": ("#7c2d12", "#ffedd5"),
                   "alum": ("#5b21b6", "#ede9fe"),
                   "local": ("#155e75", "#cffafe")}
    kfg, kbg = kind_colors.get(kind, ("#374151", "#f3f4f6"))
    contact = _html.escape(_contact(d))
    subj = f'<div style="font-size:13px;color:#374151;margin:6px 0;"><b>Subject:</b> {_html.escape(d["subject"])}</div>' if d["subject"] else ""
    body = _html.escape(d["message"]).replace("\n", "<br>")
    return f"""
<div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;margin-bottom:14px;padding:14px 18px;box-shadow:0 1px 4px rgba(0,0,0,.06);">
  <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">
    <span style="font-size:15px;font-weight:700;color:#111827;">{_html.escape(d["full_name"])}</span>
    <span style="color:#6b7280;font-size:13px;">{_html.escape(d["title"])}</span>
    {ch_badge} {_badge(kind, kfg, kbg)}
    <span style="margin-left:auto;font-size:12px;color:#6b7280;">{contact}</span>
  </div>
  {subj}
  <div style="background:#f9fafb;border-radius:8px;padding:12px 14px;margin-top:8px;
              font-size:13px;color:#1f2937;line-height:1.55;white-space:normal;">{body}</div>
</div>"""


def _write_html(drafts: list[dict], company, role):
    emails = [d for d in drafts if d["channel"] == "email"]
    lis    = [d for d in drafts if d["channel"] == "linkedin"]
    hdr = f"{_html.escape(company or 'All')} · {_html.escape(role or 'All roles')}"
    cards = ""
    if emails:
        cards += '<h2 style="font-size:15px;color:#166534;margin:18px 0 8px;">✉ Email outreach</h2>'
        cards += "".join(_card(d) for d in emails)
    if lis:
        cards += '<h2 style="font-size:15px;color:#0a66c2;margin:18px 0 8px;">in LinkedIn outreach (semi-auto, you approve each send)</h2>'
        cards += "".join(_card(d) for d in lis)
    if not drafts:
        cards = '<p style="color:#6b7280;">No drafts in the queue.</p>'

    out = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="font-family:'Segoe UI',Arial,sans-serif;background:#f3f4f6;margin:0;padding:20px;">
<div style="max-width:760px;margin:0 auto;">
  <div style="background:linear-gradient(135deg,#1e3a5f,#1d4ed8);border-radius:14px;padding:22px 26px;">
    <h1 style="color:#fff;margin:0;font-size:20px;">📨 Outreach Review Queue</h1>
    <p style="color:#bfdbfe;margin:6px 0 0;font-size:13px;">{hdr} &nbsp;·&nbsp;
       {len(emails)} email · {len(lis)} LinkedIn · <b style="color:#fff;">{len(drafts)} total</b></p>
  </div>
  <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:12px 16px;margin:12px 0;font-size:12px;color:#1e40af;">
    Edit messages in <b>outreach_review.xlsx</b> and set the <b>Action</b> column to
    <b>send</b> or <b>skip</b>, then run <code>python outreach/outreach.py --send "{_html.escape(company or '')}" "{_html.escape(role or '')}"</code>.
    Blank Action = held back (not sent).
  </div>
  {cards}
  <div style="text-align:center;padding:16px;color:#9ca3af;font-size:11px;">Job Radar · Outreach · review before sending</div>
</div></body></html>"""
    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(out)


def read_review_actions() -> list[dict]:
    """Read edited rows back from outreach_review.xlsx.
    Returns list of {log_id, subject, message, action} for rows with an action."""
    if not os.path.exists(XLSX_PATH):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[PENDING_SHEET] if PENDING_SHEET in wb.sheetnames else wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        log_id = row[_ID_COL - 1]
        # Skip company separator rows (Log ID is not an int there).
        if not isinstance(log_id, int):
            continue
        subject = row[_SUBJ_COL - 1] or ""
        message = row[_MSG_COL - 1] or ""
        action  = (row[_ACTION_COL - 1] or "").strip().lower()
        out.append({"log_id": int(log_id), "subject": subject,
                    "message": message, "action": action})
    return out


if __name__ == "__main__":
    info = write_review()
    print(f"Pending: {info['pending']} · History: {info['history']} -> {info['xlsx']}")
